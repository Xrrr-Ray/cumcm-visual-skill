#!/usr/bin/env python3
import argparse
import datetime as dt
import json
from pathlib import Path

from openpyxl import load_workbook


def normalize(value):
    if isinstance(value, (dt.datetime, dt.date, dt.time)):
        return value.isoformat()
    return value


def main():
    parser = argparse.ArgumentParser(description="把 XLSX/XLSM 工作表提取为 paper-chart 行数据")
    parser.add_argument("--input", required=True)
    parser.add_argument("--sheet")
    args = parser.parse_args()
    source = Path(args.input).resolve()
    if not source.exists():
        raise FileNotFoundError(source)
    workbook = load_workbook(source, read_only=True, data_only=True)
    if args.sheet and args.sheet not in workbook.sheetnames:
        raise ValueError(f"工作表不存在：{args.sheet}")
    sheet = workbook[args.sheet] if args.sheet else workbook[workbook.sheetnames[0]]
    values = list(sheet.iter_rows(values_only=True))
    if len(values) < 2:
        raise ValueError("工作表至少需要表头和一行数据")
    headers = [str(value).strip() if value is not None else f"column_{index + 1}" for index, value in enumerate(values[0])]
    if len(set(headers)) != len(headers):
        raise ValueError("工作表表头存在重复列名")
    rows = [
        {header: normalize(row[index] if index < len(row) else None) for index, header in enumerate(headers)}
        for row in values[1:]
        if any(value is not None and value != "" for value in row)
    ]
    # ASCII-only JSON avoids Windows console code-page corruption when Node captures stdout.
    print(json.dumps({"sheet": sheet.title, "sheets": workbook.sheetnames, "rows": rows}, ensure_ascii=True))


if __name__ == "__main__":
    main()
